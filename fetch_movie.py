# encoding:utf-8
from lxml import etree
import openpyxl
from openpyxl import styles
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
def open_file(path):
    with open(path,"r") as f:
        content=f.read()
    return content

def parse_page(page):
    movie_info={}
    movie_info_all=[]
    page_source=etree.HTML(page)
    movie_name=page_source.xpath('//*[@id="content"]/div/div[1]/table/tbody/tr/td/a/text()')
    movie_date=page_source.xpath('//*[@id="content"]/div/div[1]/table/tbody/tr/td[1]/text()')
    movie_type=page_source.xpath('//*[@id="content"]/div/div[1]/table/tbody/tr/td[3]/text()')
    movie_region=page_source.xpath('//*[@id="content"]/div/div[1]/table/tbody/tr/td[4]/text()')
    movie_wanted_num=page_source.xpath('//*[@id="content"]/div/div[1]/table/tbody/tr/td[5]/text()')
    for i in range(len(movie_name)):
        movie_info['name']=movie_name[i].strip().encode('utf-8')
        movie_info['date']=movie_date[i].strip().encode('utf-8')
        movie_info['type']=movie_type[i].strip().encode('utf-8')
        movie_info['region']=movie_region[i].strip().encode('utf-8')
        movie_info['num']=movie_wanted_num[i].strip().encode('utf-8')[:-3]
        movie_info_all.append(movie_info)
        movie_info={}
    return movie_info_all


def movie_type_filter(movie_info):
    years=[]
    diff_year={}
    diff_year['none']=[]
    for item in movie_info:
        if '年' in item['date']:
            if item['date'][0:4] in years:
                diff_year[item['date'][0:4]].append(item)
            else:
                years.append(item['date'][0:4])
                diff_year[item['date'][0:4]]=[]
                diff_year[item['date'][0:4]].append(item)
        else:
            diff_year['none'].append(item)
    return years,diff_year


def sort_by_wanted_num(years,diff_year):
    diff_year_sorted={}
    years.append('none')
    for year in years:
        diff_year_sorted[year]=sorted(diff_year[year], key=lambda each: int(each['num']), reverse=True)
    return years,diff_year_sorted


def write_info_to_xls(years,diff_year,path):
    wb=openpyxl.Workbook()
    sheet_index=0
    for year in years:
        wb.create_sheet(index=sheet_index,title=year)
        sheet_index=+1
        sheet = wb.get_sheet_by_name(year)
        sheet.cell(row=1, column=1).fill = styles.PatternFill(fill_type='solid', fgColor="e26b0a")
        for i in range(4):
            sheet.cell(row=1, column=i+2).fill = styles.PatternFill(fill_type='solid', fgColor="0070C0")
        sheet.column_dimensions['A'].width=15
        sheet.column_dimensions['B'].width =20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet['A1'] = '想看人数'
        sheet['B1'] = '上映日期'
        sheet['C1'] = '片名'
        sheet['D1'] = '类型'
        sheet['E1'] = '制片国家/地区'
        for i in range(len(diff_year[year])):
            sheet['A'+str(i+2)]=diff_year[year][i]['num']
            sheet['B'+str(i+2)]=diff_year[year][i]['date']
            sheet['C'+str(i+2)]=diff_year[year][i]['name']
            sheet['D'+str(i+2)]=diff_year[year][i]['type']
            sheet['E'+str(i+2)]=diff_year[year][i]['region']
    wb.save(path)

if __name__=='__main__':
    content=open_file('page.html')
    movie_info=parse_page(content)
    years,diff_year=movie_type_filter(movie_info)
    years,diff_year=sort_by_wanted_num(years,diff_year)
    write_info_to_xls(years,diff_year,'movie.xlsx')
